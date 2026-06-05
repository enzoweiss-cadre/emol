import json
import re
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Load data ──────────────────────────────────────────────────────────────────
with open("results/full_run_may18.json") as f:
    data = json.load(f)

results = data["results"]

# ── Load pre-enrichment for location signals ───────────────────────────────────
try:
    with open("results/pre_enrichment_results.json") as f:
        pre_raw = json.load(f)["results"]
    PRE = {r["domain"]: r for r in pre_raw}
except Exception:
    PRE = {}

# ── Load IP geolocation results ────────────────────────────────────────────────
try:
    with open("results/ip_location_results.json") as f:
        IP_GEO = json.load(f)
except Exception:
    IP_GEO = {}

# ── Load usage time windows ────────────────────────────────────────────────────
try:
    with open("results/usage_windows.json") as f:
        USAGE_WINDOWS = json.load(f)
except Exception:
    USAGE_WINDOWS = {}

# ── Load Sonar research location ───────────────────────────────────────────────
try:
    with open("results/research_location_results.json") as f:
        RESEARCH_LOC = json.load(f)
except Exception:
    RESEARCH_LOC = {}

# Country-code TLD → country name
TLD_COUNTRY = {
    "cn": "China", "de": "Germany", "fr": "France", "jp": "Japan",
    "uk": "United Kingdom", "co.uk": "United Kingdom", "in": "India",
    "ch": "Switzerland", "kr": "South Korea", "co.kr": "South Korea",
    "se": "Sweden", "dk": "Denmark", "nl": "Netherlands", "be": "Belgium",
    "es": "Spain", "it": "Italy", "au": "Australia", "ca": "Canada",
    "il": "Israel", "sg": "Singapore", "tw": "Taiwan", "br": "Brazil",
    "ru": "Russia", "at": "Austria", "no": "Norway", "fi": "Finland",
    "pl": "Poland", "cz": "Czech Republic", "hu": "Hungary",
    "ie": "Ireland", "pt": "Portugal", "gr": "Greece", "ro": "Romania",
}

EU_COUNTRIES = {
    "Germany", "France", "Switzerland", "Sweden", "Denmark", "Netherlands",
    "Belgium", "Spain", "Italy", "Austria", "Norway", "Finland", "Poland",
    "Czech Republic", "Hungary", "Ireland", "Portugal", "Greece", "Romania",
    "United Kingdom",
}

# City → country mapping for keyword extraction
CITY_COUNTRY = {
    "boston": "United States", "cambridge": "United States",
    "san francisco": "United States", "san diego": "United States",
    "new york": "United States", "seattle": "United States",
    "chicago": "United States", "los angeles": "United States",
    "houston": "United States", "raleigh": "United States",
    "philadelphia": "United States", "atlanta": "United States",
    "london": "United Kingdom", "zurich": "Switzerland",
    "basel": "Switzerland", "munich": "Germany", "berlin": "Germany",
    "frankfurt": "Germany", "hamburg": "Germany",
    "paris": "France", "lyon": "France",
    "tokyo": "Japan", "osaka": "Japan",
    "shanghai": "China", "beijing": "China", "shenzhen": "China",
    "guangzhou": "China", "suzhou": "China", "wuhan": "China",
    "bangalore": "India", "hyderabad": "India", "mumbai": "India",
    "pune": "India", "chennai": "India",
    "singapore": "Singapore", "toronto": "Canada",
    "tel aviv": "Israel", "rehovot": "Israel",
    "amsterdam": "Netherlands", "stockholm": "Sweden",
    "copenhagen": "Denmark", "oslo": "Norway", "helsinki": "Finland",
}

_CITY_RE = re.compile(
    r'\b(' + '|'.join(re.escape(c) for c in CITY_COUNTRY) + r')\b',
    re.IGNORECASE
)
_US_STATE_RE = re.compile(
    r'\b(California|Massachusetts|New Jersey|Connecticut|Maryland|'
    r'North Carolina|Texas|Illinois|Washington|Pennsylvania|'
    r'New York State|Colorado|Georgia|Indiana|Ohio|Florida|Virginia|'
    r'Missouri|Tennessee)\b',
    re.IGNORECASE
)

def infer_location(domain: str) -> tuple[str, str, str]:
    """Returns (country, city, source)."""
    p = PRE.get(domain, {})

    # Registry sources — highest confidence
    if p.get("sirene_found"):
        return "France", p.get("sirene_city", ""), "registry"
    if p.get("zefix_found"):
        return "Switzerland", "", "registry"
    if p.get("opendart_registry") == "registered":
        return "South Korea", "", "registry"

    # Country-code TLD
    for tld, country in TLD_COUNTRY.items():
        if domain.endswith("." + tld):
            return country, "", "tld"

    # IP geolocation — actual user location
    ip_data = IP_GEO.get(domain)
    if ip_data and ip_data.get("country"):
        return ip_data["country"], ip_data.get("city", ""), "ip_geo"

    # Keyword search in scraped/web text
    text = " ".join(filter(None, [
        p.get("sig_web_context") or "",
        p.get("site_summary_text") or "",
    ]))
    if text:
        m = _CITY_RE.search(text)
        if m:
            return CITY_COUNTRY[m.group(1).lower()], m.group(1).title(), "keyword"
        if _US_STATE_RE.search(text):
            return "United States", "", "keyword"

    return "", "", ""


def country_to_tier(country: str) -> str:
    if not country:
        return "Unknown"
    if country == "United States":
        return "US"
    if country in EU_COUNTRIES:
        return "EU"
    if country == "China":
        return "China"
    if country == "India":
        return "India"
    return "Other"


# Enforce qualification_status from final_icp_score (override agent text)
for r in results:
    score = r.get("final_icp_score") or 0
    if score >= 80:
        r["qualification_status"] = "HIGH FIT"
    elif score >= 60:
        r["qualification_status"] = "MEDIUM FIT"
    elif score >= 40:
        r["qualification_status"] = "LOW FIT"
    else:
        r["qualification_status"] = "NO FIT"

# Pre-compute location for all results
for r in results:
    country, city, source = infer_location(r.get("domain", ""))
    r["_hq_country"]    = country
    r["_hq_city"]       = city
    r["_loc_source"]    = source
    r["_geo_tier"]      = country_to_tier(country)

# ── Colour palette ─────────────────────────────────────────────────────────────
C_HIGH    = "FF4CAF50"   # green
C_MEDIUM  = "FFFFC107"   # amber
C_LOW     = "FFFF9800"   # orange
C_NO      = "FFF44336"   # red
C_HEADER  = "FF1A237E"   # deep indigo
C_SUBHDR  = "FF283593"   # lighter indigo
C_SM      = "FFE3F2FD"   # light blue bg
C_SF      = "FFF3E5F5"   # light purple bg
C_CQ      = "FFE8F5E9"   # light green bg
C_SI      = "FFFFF3E0"   # light orange bg
C_WHITE   = "FFFFFFFF"
C_GRAY    = "FFF5F5F5"

def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="FFE0E0E0")
    return Border(left=s, right=s, top=s, bottom=s)

def score_fill(score):
    if score is None:
        return fill(C_GRAY)
    if score >= 80:
        return fill(C_HIGH)
    if score >= 60:
        return fill(C_MEDIUM)
    if score >= 40:
        return fill(C_LOW)
    return fill(C_NO)

def score_font(score):
    if score is None:
        return Font(bold=False, color="FF9E9E9E")
    if score >= 80:
        return Font(bold=True, color="FF1B5E20")
    if score >= 60:
        return Font(bold=True, color="FF4E342E")
    if score >= 40:
        return Font(bold=True, color="FFE65100")
    return Font(bold=True, color="FFB71C1C")

def status_fill(status):
    m = {"HIGH FIT": C_HIGH, "MEDIUM FIT": C_MEDIUM, "LOW FIT": C_LOW, "NO FIT": C_NO}
    return fill(m.get(status, C_GRAY))

# ── Column definitions ─────────────────────────────────────────────────────────
# Each entry: (header, field_or_fn, width, bg_hint)
C_LOC  = "FFE8EAF6"   # light lavender bg for location cols
C_USE  = "FFFCE4EC"   # light pink bg for usage cols

COLS = [
    # Identity
    ("Domain",               lambda r: r.get("domain", ""),                         28, None),
    ("Company Type",         lambda r: r.get("company_type", ""),                   16, None),
    ("Status",               lambda r: r.get("qualification_status", ""),           14, None),

    # Total score
    ("TOTAL SCORE",          lambda r: r.get("final_icp_score"),                    13, None),

    # ── Location ──────────────────────────────────────────────────────────────
    ("Geo Tier",             lambda r: r.get("_geo_tier", ""),                      12, C_LOC),
    ("Country (IP/Reg)",     lambda r: r.get("_hq_country", ""),                   20, C_LOC),
    ("City (IP/Reg)",        lambda r: r.get("_hq_city", ""),                      18, C_LOC),
    ("Research Country",     lambda r: RESEARCH_LOC.get(r.get("domain",""),{}).get("country",""),  20, C_LOC),
    ("Research City",        lambda r: RESEARCH_LOC.get(r.get("domain",""),{}).get("city",""),     18, C_LOC),
    ("Research Detail",      lambda r: RESEARCH_LOC.get(r.get("domain",""),{}).get("raw",""),      50, C_LOC),

    # ── Usage ─────────────────────────────────────────────────────────────────
    ("Users",                lambda r: r.get("source_user_count"),                  10, C_USE),
    ("Total Sessions",       lambda r: r.get("source_total_sessions"),              14, C_USE),
    ("Last Active",          lambda r: (r.get("source_last_active") or "")[:10],   14, C_USE),
    ("Users (30d)",          lambda r: USAGE_WINDOWS.get(r.get("domain",""),{}).get("users_30d"),  11, C_USE),
    ("Users (6mo)",          lambda r: USAGE_WINDOWS.get(r.get("domain",""),{}).get("users_6mo"),  11, C_USE),
    ("Users (12mo)",         lambda r: USAGE_WINDOWS.get(r.get("domain",""),{}).get("users_12mo"), 11, C_USE),

    # ── SM ────────────────────────────────────────────────────────────────────
    ("SM Score",             lambda r: r.get("small_molecule_score"),               10, C_SM),
    ("SM · Active MedChem",  lambda r: r.get("sm_active_med_chem", ""),            36, C_SM),
    ("SM · CADD",            lambda r: r.get("sm_cadd_capabilities", ""),          36, C_SM),
    ("SM · Hit-to-Lead",     lambda r: r.get("sm_hit_to_lead", ""),               36, C_SM),
    ("SM · Virtual Screen",  lambda r: r.get("sm_virtual_screening", ""),         36, C_SM),
    ("SM · Basis",           lambda r: r.get("sm_scoring_basis", ""),             28, C_SM),

    # ── SF ────────────────────────────────────────────────────────────────────
    ("SF Score",             lambda r: r.get("stage_fit_score"),                   10, C_SF),
    ("SF · Stage Type",      lambda r: r.get("sf_stage_type", ""),                18, C_SF),
    ("SF · Funding Stage",   lambda r: r.get("sf_funding_stage", ""),             22, C_SF),
    ("SF · Employee Range",  lambda r: r.get("sf_employee_range", ""),            20, C_SF),
    ("SF · Is Public",       lambda r: r.get("sf_is_public", ""),                 30, C_SF),
    ("SF · Basis",           lambda r: r.get("sf_scoring_basis", ""),             36, C_SF),

    # ── CQ ────────────────────────────────────────────────────────────────────
    ("CQ Score",             lambda r: r.get("contact_quality_score"),             10, C_CQ),
    ("CQ · Has Chem Team",   lambda r: r.get("cq_has_chemistry_team", ""),        36, C_CQ),
    ("CQ · Highest Role",    lambda r: r.get("cq_highest_role", ""),              28, C_CQ),
    ("CQ · Team Size Est.",  lambda r: r.get("cq_team_size_estimate", ""),        22, C_CQ),
    ("CQ · Basis",           lambda r: r.get("cq_scoring_basis", ""),             28, C_CQ),

    # ── SI ────────────────────────────────────────────────────────────────────
    ("SI Score",             lambda r: r.get("strategic_score"),                   10, C_SI),
    ("SI · Funding Amount",  lambda r: r.get("si_funding_amount", ""),            28, C_SI),
    ("SI · Patent Filings",  lambda r: r.get("si_patent_filings", ""),            36, C_SI),
    ("SI · Partnerships",    lambda r: r.get("si_pharma_partnerships", ""),       36, C_SI),
    ("SI · Recent Hires",    lambda r: r.get("si_recent_chem_hires", ""),         36, C_SI),
    ("SI · Recent Funding",  lambda r: r.get("si_recent_funding", ""),            36, C_SI),
    ("SI · Publications",    lambda r: r.get("si_recent_publications", ""),       36, C_SI),
    ("SI · Basis",           lambda r: r.get("si_scoring_basis", ""),             28, C_SI),

    # Extra context
    ("DQ Reason",            lambda r: r.get("disqualify_reason", "") or "",      50, None),
    ("Key Evidence",         lambda r: " | ".join(r.get("key_evidence", [])),     70, None),
    ("Agent Reasoning",      lambda r: r.get("reasoning", ""),                    80, None),
]

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: All companies ─────────────────────────────────────────────────────
ws = wb.active
ws.title = "All Companies"

# Freeze top row
ws.freeze_panes = "A2"

# Header row
header_font  = Font(bold=True, color="FFFFFFFF", size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for ci, (hdr, _, width, bg) in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=ci, value=hdr)
    cell.font = header_font
    cell.alignment = header_align
    cell.fill = fill(C_SUBHDR) if bg else fill(C_HEADER)
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(ci)].width = width

ws.row_dimensions[1].height = 36

# Sort by score desc
sorted_results = sorted(results, key=lambda r: r.get("final_icp_score") or 0, reverse=True)

wrap = Alignment(vertical="top", wrap_text=True)
center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

for ri, rec in enumerate(sorted_results, 2):
    row_bg = C_WHITE if ri % 2 == 0 else C_GRAY
    for ci, (hdr, getter, _, bg) in enumerate(COLS, 1):
        val  = getter(rec)
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.border = thin_border()
        cell.alignment = center_wrap if ci <= 5 else wrap

        # Score columns get colour-coded fills
        if hdr in ("TOTAL SCORE", "SM Score", "SF Score", "CQ Score", "SI Score"):
            cell.fill = score_fill(val)
            cell.font = score_font(val)
        elif hdr == "Status":
            cell.fill = status_fill(val)
            cell.font = Font(bold=True, size=9)
        elif hdr == "Geo Tier":
            tier_colors = {
                "US":      "FFBBDEFB",   # blue
                "EU":      "FFC8E6C9",   # green
                "China":   "FFFFCDD2",   # red
                "India":   "FFFFE0B2",   # orange
                "Other":   "FFF5F5F5",   # gray
                "Unknown": "FFFFFFFF",
            }
            cell.fill = fill(tier_colors.get(val, "FFFFFFFF"))
            cell.font = Font(bold=bool(val and val != "Unknown"), size=9)
        elif bg:
            cell.fill = fill(bg)
        else:
            cell.fill = fill(row_bg)

    ws.row_dimensions[ri].height = 60

# Auto-filter on header row
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ── Sheet 2: HIGH FIT only ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("HIGH FIT")
ws2.freeze_panes = "A2"

for ci, (hdr, _, width, bg) in enumerate(COLS, 1):
    cell = ws2.cell(row=1, column=ci, value=hdr)
    cell.font = header_font
    cell.alignment = header_align
    cell.fill = fill("FF2E7D32") if bg else fill("FF1B5E20")
    cell.border = thin_border()
    ws2.column_dimensions[get_column_letter(ci)].width = width
ws2.row_dimensions[1].height = 36

high_fit = [r for r in sorted_results if r.get("qualification_status") == "HIGH FIT"]
for ri, rec in enumerate(high_fit, 2):
    row_bg = C_WHITE if ri % 2 == 0 else "FFE8F5E9"
    for ci, (hdr, getter, _, bg) in enumerate(COLS, 1):
        val  = getter(rec)
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.border = thin_border()
        cell.alignment = center_wrap if ci <= 5 else wrap
        if hdr in ("TOTAL SCORE", "SM Score", "SF Score", "CQ Score", "SI Score"):
            cell.fill = score_fill(val)
            cell.font = score_font(val)
        elif hdr == "Status":
            cell.fill = status_fill(val)
            cell.font = Font(bold=True, size=9)
        elif hdr == "Geo Tier":
            tier_colors = {
                "US":      "FFBBDEFB",
                "EU":      "FFC8E6C9",
                "China":   "FFFFCDD2",
                "India":   "FFFFE0B2",
                "Other":   "FFF5F5F5",
                "Unknown": "FFFFFFFF",
            }
            cell.fill = fill(tier_colors.get(val, "FFFFFFFF"))
            cell.font = Font(bold=bool(val and val != "Unknown"), size=9)
        elif bg:
            cell.fill = fill(bg)
        else:
            cell.fill = fill(row_bg)
    ws2.row_dimensions[ri].height = 60

ws2.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ── Sheet 3: Score distribution summary ───────────────────────────────────────
ws3 = wb.create_sheet("Summary")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 12

summary_data = [
    ("Tier",          "Count", "% of Total"),
    ("HIGH FIT",      None,    None),
    ("MEDIUM FIT",    None,    None),
    ("LOW FIT",       None,    None),
    ("NO FIT",        None,    None),
    ("TOTAL",         None,    None),
]

tier_map = {}
for r in results:
    t = r.get("qualification_status", "Unknown")
    tier_map[t] = tier_map.get(t, 0) + 1
total = len(results)

tiers = ["HIGH FIT", "MEDIUM FIT", "LOW FIT", "NO FIT"]
tier_colors = [C_HIGH, C_MEDIUM, C_LOW, C_NO]

# Write header
for ci, h in enumerate(["Tier", "Count", "% of Total"], 1):
    c = ws3.cell(row=1, column=ci, value=h)
    c.font = Font(bold=True, color="FFFFFFFF")
    c.fill = fill(C_HEADER)
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()

for ri, (tier, color) in enumerate(zip(tiers, tier_colors), 2):
    count = tier_map.get(tier, 0)
    pct   = f"{count/total*100:.1f}%"
    ws3.cell(row=ri, column=1, value=tier).fill = fill(color)
    ws3.cell(row=ri, column=2, value=count).fill = fill(color)
    ws3.cell(row=ri, column=3, value=pct).fill = fill(color)
    for ci in range(1, 4):
        ws3.cell(row=ri, column=ci).border = thin_border()
        ws3.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
        ws3.cell(row=ri, column=ci).font = Font(bold=True)

# Total row
tr = len(tiers) + 2
ws3.cell(row=tr, column=1, value="TOTAL").fill = fill(C_HEADER)
ws3.cell(row=tr, column=2, value=total).fill = fill(C_HEADER)
ws3.cell(row=tr, column=3, value="100%").fill = fill(C_HEADER)
for ci in range(1, 4):
    ws3.cell(row=tr, column=ci).font = Font(bold=True, color="FFFFFFFF")
    ws3.cell(row=tr, column=ci).border = thin_border()
    ws3.cell(row=tr, column=ci).alignment = Alignment(horizontal="center")

# ── Sheet 4: Location & Usage breakdown ───────────────────────────────────────
ws4 = wb.create_sheet("Location & Usage")
ws4.column_dimensions["A"].width = 14
ws4.column_dimensions["B"].width = 10
ws4.column_dimensions["C"].width = 10
ws4.column_dimensions["D"].width = 16
ws4.column_dimensions["E"].width = 16
ws4.column_dimensions["F"].width = 16

loc_headers = ["Geo Tier", "Companies", "HIGH FIT", "Avg Sessions", "Total Sessions", "Avg Users"]
for ci, h in enumerate(loc_headers, 1):
    c = ws4.cell(row=1, column=ci, value=h)
    c.font = Font(bold=True, color="FFFFFFFF")
    c.fill = fill(C_HEADER)
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()

tier_order = ["US", "EU", "India", "China", "Other", "Unknown"]
tier_bg = {
    "US":      "FFBBDEFB", "EU":   "FFC8E6C9",
    "China":   "FFFFCDD2", "India": "FFFFE0B2",
    "Other":   "FFF5F5F5", "Unknown": "FFFFFFFF",
}

for ri, tier in enumerate(tier_order, 2):
    bucket = [r for r in results if r.get("_geo_tier") == tier]
    high   = sum(1 for r in bucket if r.get("qualification_status") == "HIGH FIT")
    sessions = [r.get("source_total_sessions") or 0 for r in bucket]
    users    = [r.get("source_user_count") or 0 for r in bucket]
    avg_sess = round(sum(sessions) / len(sessions), 1) if sessions else 0
    tot_sess = sum(sessions)
    avg_usr  = round(sum(users) / len(users), 1) if users else 0

    row_vals = [tier, len(bucket), high, avg_sess, tot_sess, avg_usr]
    bg = tier_bg.get(tier, "FFFFFFFF")
    for ci, v in enumerate(row_vals, 1):
        c = ws4.cell(row=ri, column=ci, value=v)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=(ci == 1))

# ── Save ───────────────────────────────────────────────────────────────────────
out = "results/emolecules_scores_may18.xlsx"
wb.save(out)
print(f"Saved → {out}")
print(f"  All companies : {len(sorted_results)}")
print(f"  HIGH FIT sheet: {len(high_fit)}")

# Print location coverage stats
from collections import Counter
tier_counts = Counter(r.get("_geo_tier") for r in results)
print("\nLocation coverage:")
for tier in tier_order:
    print(f"  {tier:10}: {tier_counts.get(tier, 0)}")
